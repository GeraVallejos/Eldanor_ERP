import csv
import io

from apps.core.exceptions import BusinessRuleError

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency guard
    load_workbook = None


MAX_IMPORT_FILE_BYTES = 2 * 1024 * 1024
MAX_IMPORT_ROWS = 2000


def _normalize_header(value):
    return str(value or "").strip().lower()


def _cell_to_string(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _validate_required_headers(fieldnames, required_headers):
    normalized_headers = {_normalize_header(header) for header in fieldnames}
    required_set = {_normalize_header(header) for header in required_headers}
    missing_headers = sorted(required_set - normalized_headers)
    if missing_headers:
        raise BusinessRuleError(
            "Faltan columnas obligatorias en el archivo.",
            error_code="BULK_IMPORT_MISSING_REQUIRED_COLUMNS",
            meta={"missing_columns": missing_headers},
        )


def _parse_rows_from_csv(text, required_headers, *, max_rows):
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise BusinessRuleError(
            "El archivo CSV no contiene cabeceras.",
            error_code="BULK_IMPORT_MISSING_HEADERS",
        )

    _validate_required_headers(reader.fieldnames, required_headers)

    rows = []
    for line_number, row in enumerate(reader, start=2):
        normalized_row = {
            _normalize_header(key): _cell_to_string(value)
            for key, value in (row or {}).items()
            if key is not None
        }

        if not any(normalized_row.values()):
            continue

        rows.append((line_number, normalized_row))
        if len(rows) > max_rows:
            raise BusinessRuleError(
                f"La carga masiva admite hasta {max_rows} filas por archivo.",
                error_code="BULK_IMPORT_TOO_MANY_ROWS",
            )

    return rows


def _parse_rows_from_xlsx(raw_content, required_headers, *, max_rows):
    if load_workbook is None:
        raise BusinessRuleError(
            "No se pudo procesar XLSX porque falta la dependencia openpyxl.",
            error_code="BULK_IMPORT_XLSX_DEPENDENCY_MISSING",
        )

    try:
        workbook = load_workbook(filename=io.BytesIO(raw_content), read_only=True, data_only=True)
    except Exception as exc:
        raise BusinessRuleError(
            "El archivo XLSX no es valido o esta corrupto.",
            error_code="BULK_IMPORT_INVALID_XLSX",
        ) from exc

    try:
        sheet = workbook.active
        row_iterator = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(row_iterator)
        except StopIteration as exc:
            raise BusinessRuleError(
                "El archivo XLSX no contiene cabeceras.",
                error_code="BULK_IMPORT_MISSING_HEADERS",
            ) from exc

        headers = [_normalize_header(value) for value in raw_headers]
        if not any(headers):
            raise BusinessRuleError(
                "El archivo XLSX no contiene cabeceras.",
                error_code="BULK_IMPORT_MISSING_HEADERS",
            )

        _validate_required_headers(headers, required_headers)

        rows = []
        for line_number, values in enumerate(row_iterator, start=2):
            normalized_row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                cell_value = values[index] if index < len(values) else None
                normalized_row[header] = _cell_to_string(cell_value)

            if not any(normalized_row.values()):
                continue

            rows.append((line_number, normalized_row))
            if len(rows) > max_rows:
                raise BusinessRuleError(
                    f"La carga masiva admite hasta {max_rows} filas por archivo.",
                    error_code="BULK_IMPORT_TOO_MANY_ROWS",
                )

        return rows
    finally:
        workbook.close()


def parse_csv_upload(uploaded_file, required_headers, *, max_rows=MAX_IMPORT_ROWS, max_bytes=MAX_IMPORT_FILE_BYTES):
    if not uploaded_file:
        raise BusinessRuleError(
            "Debe adjuntar un archivo CSV o XLSX para la carga masiva.",
            error_code="BULK_IMPORT_FILE_REQUIRED",
        )

    filename = str(getattr(uploaded_file, "name", "") or "").lower()
    is_csv = filename.endswith(".csv")
    is_xlsx = filename.endswith(".xlsx")
    if not is_csv and not is_xlsx:
        raise BusinessRuleError(
            "Formato de archivo no soportado. Solo se permite CSV o XLSX.",
            error_code="BULK_IMPORT_INVALID_FILE_TYPE",
        )

    file_size = int(getattr(uploaded_file, "size", 0) or 0)
    if file_size > max_bytes:
        raise BusinessRuleError(
            f"Archivo demasiado grande. Maximo permitido: {max_bytes // (1024 * 1024)} MB.",
            error_code="BULK_IMPORT_FILE_TOO_LARGE",
        )

    raw_content = uploaded_file.read()
    if len(raw_content) > max_bytes:
        raise BusinessRuleError(
            f"Archivo demasiado grande. Maximo permitido: {max_bytes // (1024 * 1024)} MB.",
            error_code="BULK_IMPORT_FILE_TOO_LARGE",
        )

    if is_csv:
        try:
            text = raw_content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise BusinessRuleError(
                "No se pudo leer el CSV. Use codificacion UTF-8.",
                error_code="BULK_IMPORT_INVALID_ENCODING",
            ) from exc

        if not text.strip():
            raise BusinessRuleError(
                "El archivo CSV no contiene datos.",
                error_code="BULK_IMPORT_EMPTY_FILE",
            )

        return _parse_rows_from_csv(text, required_headers, max_rows=max_rows)

    rows = _parse_rows_from_xlsx(raw_content, required_headers, max_rows=max_rows)
    if not rows:
        raise BusinessRuleError(
            "El archivo XLSX no contiene datos.",
            error_code="BULK_IMPORT_EMPTY_FILE",
        )

    return rows
