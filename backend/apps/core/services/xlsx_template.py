import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def build_xlsx_template(*, headers, sample_row=None, instructions=None, sheet_name="Plantilla"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[cell.column_letter].width = max(14, min(42, len(str(header)) + 4))

    if sample_row:
        padded_sample = list(sample_row)[: len(headers)]
        while len(padded_sample) < len(headers):
            padded_sample.append("")
        sheet.append(padded_sample)

    if instructions:
        help_sheet = workbook.create_sheet(title="Instrucciones")
        help_sheet.append(["Indicacion"])
        help_sheet["A1"].font = Font(bold=True)
        help_sheet.column_dimensions["A"].width = 120
        for item in instructions:
            help_sheet.append([str(item or "")])

    # Remove sheet protection from all sheets
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        ws.protection.sheet = False

    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()
